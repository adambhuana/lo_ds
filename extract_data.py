import openpyxl

wb = openpyxl.load_workbook('learning_outcome_ds.xlsx')
ws = wb.active

print(f"Sheet: {ws.title}")
print(f"Max rows: {ws.max_row}")
print(f"Max cols: {ws.max_column}")
print("=" * 120)

for r in range(1, ws.max_row + 1):
    row_data = []
    for c in range(1, ws.max_column + 1):
        val = ws.cell(r, c).value
        if val is not None:
            s = str(val)[:100]
        else:
            s = ""
        row_data.append(s)
    print(f"Row {r:3d}: | {'  |  '.join(row_data)}")
